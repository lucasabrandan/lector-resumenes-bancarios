"""
Tests del módulo de reportes Ley 25.413.

Cubre:
- Cálculo de totales mensuales (bruto, neto, devoluciones).
- Exportación a XLSX.
- Caso sin datos.
"""

from datetime import date
from decimal import Decimal

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.db.base import Base
from app.db.models import MovimientoDB
from app.reports.ley_25413 import (
    TotalMensualLey25413,
    generar_reporte_ley_25413,
    resumen_general,
    exportar_reporte_xlsx,
)


@pytest.fixture
def db_session():
    """Sesión de DB en memoria para tests."""
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


def _crear_movimiento(
    fecha: date,
    importe: float,
    tipo: str,
    signo: str = "DEBITO",
    concepto: str = "Test",
) -> MovimientoDB:
    return MovimientoDB(
        banco="SUPERVIELLE",
        cuenta="001",
        fecha=fecha,
        concepto=concepto,
        importe=importe,
        signo=signo,
        moneda="ARS",
        tipo=tipo,
    )


class TestTotalMensualLey25413:
    """Tests del dataclass TotalMensualLey25413."""

    def test_total_bruto(self):
        t = TotalMensualLey25413(
            anio=2024, mes=1,
            impuesto_sobre_debitos=Decimal("1000.00"),
            impuesto_sobre_creditos=Decimal("300.00"),
        )
        assert t.total_bruto == Decimal("1300.00")

    def test_total_neto_sin_devoluciones(self):
        t = TotalMensualLey25413(
            anio=2024, mes=1,
            impuesto_sobre_debitos=Decimal("1000.00"),
            impuesto_sobre_creditos=Decimal("300.00"),
        )
        assert t.total_neto == Decimal("1300.00")
        assert t.neto_sobre_debitos == Decimal("1000.00")

    def test_total_neto_con_devoluciones(self):
        t = TotalMensualLey25413(
            anio=2024, mes=1,
            impuesto_sobre_debitos=Decimal("1000.00"),
            impuesto_sobre_creditos=Decimal("300.00"),
            devoluciones_debitos=Decimal("150.00"),
        )
        assert t.neto_sobre_debitos == Decimal("850.00")
        assert t.total_neto == Decimal("1150.00")

    def test_periodo(self):
        t = TotalMensualLey25413(
            anio=2024, mes=3,
            impuesto_sobre_debitos=Decimal("0"),
            impuesto_sobre_creditos=Decimal("0"),
        )
        assert t.periodo == "Marzo 2024"
        assert t.mes_nombre == "Marzo"

    def test_mes_nombre_todos_los_meses(self):
        nombres = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
        ]
        for i, nombre in enumerate(nombres, 1):
            t = TotalMensualLey25413(
                anio=2024, mes=i,
                impuesto_sobre_debitos=Decimal("0"),
                impuesto_sobre_creditos=Decimal("0"),
            )
            assert t.mes_nombre == nombre


class TestGenerarReporteLey25413:
    """Tests de la query que genera el reporte desde la DB."""

    def test_sin_datos(self, db_session):
        resultado = generar_reporte_ley_25413(db_session)
        assert resultado == []

    def test_un_mes_simple(self, db_session):
        db_session.add_all([
            _crear_movimiento(date(2024, 1, 15), 1000.50, "IMPUESTO_LEY_25413_SOBRE_DEBITOS"),
            _crear_movimiento(date(2024, 1, 20), 500.25, "IMPUESTO_LEY_25413_SOBRE_DEBITOS"),
            _crear_movimiento(date(2024, 1, 25), 200.00, "IMPUESTO_LEY_25413_SOBRE_CREDITOS"),
        ])
        db_session.commit()

        resultado = generar_reporte_ley_25413(db_session)
        assert len(resultado) == 1
        r = resultado[0]
        assert r.anio == 2024
        assert r.mes == 1
        assert r.impuesto_sobre_debitos == Decimal("1500.75")
        assert r.impuesto_sobre_creditos == Decimal("200.00")
        assert r.total_bruto == Decimal("1700.75")

    def test_varios_meses_ordenados(self, db_session):
        db_session.add_all([
            _crear_movimiento(date(2024, 3, 10), 100.00, "IMPUESTO_LEY_25413_SOBRE_DEBITOS"),
            _crear_movimiento(date(2024, 1, 10), 200.00, "IMPUESTO_LEY_25413_SOBRE_DEBITOS"),
            _crear_movimiento(date(2024, 2, 10), 300.00, "IMPUESTO_LEY_25413_SOBRE_DEBITOS"),
        ])
        db_session.commit()

        resultado = generar_reporte_ley_25413(db_session)
        assert len(resultado) == 3
        assert resultado[0].mes == 1
        assert resultado[1].mes == 2
        assert resultado[2].mes == 3

    def test_con_devoluciones(self, db_session):
        db_session.add_all([
            _crear_movimiento(date(2024, 1, 15), 1000.00, "IMPUESTO_LEY_25413_SOBRE_DEBITOS"),
            _crear_movimiento(date(2024, 1, 20), 300.00, "IMPUESTO_LEY_25413_SOBRE_CREDITOS"),
            _crear_movimiento(
                date(2024, 1, 25), 60.00, "DEVOLUCION",
                signo="CREDITO",
                concepto="Devolución Imp. Débitos",
            ),
        ])
        db_session.commit()

        resultado = generar_reporte_ley_25413(db_session)
        assert len(resultado) == 1
        r = resultado[0]
        assert r.impuesto_sobre_debitos == Decimal("1000.00")
        assert r.devoluciones_debitos == Decimal("60.00")
        assert r.neto_sobre_debitos == Decimal("940.00")
        assert r.total_neto == Decimal("1240.00")

    def test_devoluciones_no_relacionadas_no_se_descuentan(self, db_session):
        """Una devolución que NO es de impuesto a débitos no debe restarse."""
        db_session.add_all([
            _crear_movimiento(date(2024, 1, 15), 1000.00, "IMPUESTO_LEY_25413_SOBRE_DEBITOS"),
            _crear_movimiento(
                date(2024, 1, 25), 200.00, "DEVOLUCION",
                signo="CREDITO",
                concepto="Devolución Extracción ATM",
            ),
        ])
        db_session.commit()

        resultado = generar_reporte_ley_25413(db_session)
        assert len(resultado) == 1
        r = resultado[0]
        assert r.devoluciones_debitos == Decimal("0.00")
        assert r.neto_sobre_debitos == Decimal("1000.00")

    def test_ignora_otros_tipos(self, db_session):
        """Movimientos que no son impuesto ni devolución no aparecen."""
        db_session.add_all([
            _crear_movimiento(date(2024, 1, 10), 5000.00, "COMISION"),
            _crear_movimiento(date(2024, 1, 11), 1000.00, "IVA"),
            _crear_movimiento(date(2024, 1, 15), 100.00, "IMPUESTO_LEY_25413_SOBRE_DEBITOS"),
        ])
        db_session.commit()

        resultado = generar_reporte_ley_25413(db_session)
        assert len(resultado) == 1
        assert resultado[0].impuesto_sobre_debitos == Decimal("100.00")


class TestResumenGeneral:
    """Tests de resumen_general."""

    def test_sin_datos(self, db_session):
        resultado = resumen_general(db_session)
        assert resultado["total_movimientos"] == 0
        assert resultado["total_debitos"] == Decimal("0")

    def test_con_datos(self, db_session):
        db_session.add_all([
            _crear_movimiento(date(2024, 1, 10), 100.00, "COMISION", signo="DEBITO"),
            _crear_movimiento(date(2024, 1, 15), 200.00, "COMISION", signo="DEBITO"),
            _crear_movimiento(date(2024, 1, 20), 500.00, "TRANSFERENCIA_RECIBIDA", signo="CREDITO"),
        ])
        db_session.commit()

        resultado = resumen_general(db_session)
        assert resultado["total_movimientos"] == 3
        assert resultado["total_debitos"] == Decimal("300.00")
        assert resultado["cant_debitos"] == 2
        assert resultado["total_creditos"] == Decimal("500.00")
        assert resultado["cant_creditos"] == 1


class TestExportarXLSX:
    """Tests de la exportación a XLSX."""

    def test_genera_archivo_valido(self):
        from openpyxl import load_workbook

        reporte = [
            TotalMensualLey25413(
                anio=2024, mes=1,
                impuesto_sobre_debitos=Decimal("1000.00"),
                impuesto_sobre_creditos=Decimal("300.00"),
                devoluciones_debitos=Decimal("50.00"),
            ),
            TotalMensualLey25413(
                anio=2024, mes=2,
                impuesto_sobre_debitos=Decimal("800.00"),
                impuesto_sobre_creditos=Decimal("250.00"),
            ),
        ]

        xlsx = exportar_reporte_xlsx(reporte)
        wb = load_workbook(xlsx)
        ws = wb.active

        assert ws.title == "Ley 25413"
        # Título
        assert "Ley 25.413" in ws["A1"].value
        # Headers en fila 4
        assert ws.cell(row=4, column=1).value == "Período"
        # Datos fila 5 (Enero 2024)
        assert ws.cell(row=5, column=1).value == "Enero 2024"
        assert ws.cell(row=5, column=2).value == 1000.00
        assert ws.cell(row=5, column=3).value == 50.00
        assert ws.cell(row=5, column=4).value == 950.00  # neto
        # Datos fila 6 (Febrero 2024)
        assert ws.cell(row=6, column=1).value == "Febrero 2024"
        assert ws.cell(row=6, column=3).value == 0.00  # sin devoluciones
        # Fila de totales (fila 7)
        assert ws.cell(row=7, column=1).value == "TOTAL"
        assert ws.cell(row=7, column=2).value == 1800.00  # total bruto
        assert ws.cell(row=7, column=6).value == 2300.00  # total neto

    def test_reporte_vacio(self):
        from openpyxl import load_workbook

        xlsx = exportar_reporte_xlsx([])
        wb = load_workbook(xlsx)
        ws = wb.active

        assert ws.cell(row=4, column=1).value == "Período"
        # Fila 5 es la de totales (sin datos)
        assert ws.cell(row=5, column=1).value == "TOTAL"
        assert ws.cell(row=5, column=2).value == 0.0
