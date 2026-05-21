"""
Parser de archivos "Mis Comprobantes Emitidos" exportados desde ARCA.

Soporta CSV (separado por ;) y Excel (.xlsx/.xls).
Columnas esperadas del archivo ARCA:
  Fecha | Tipo | Punto de Venta | Número Desde | Número Hasta |
  Cód. Autorización | Tipo Doc. Receptor | Nro. Doc. Receptor |
  Denominación Receptor | Tipo Cambio | Moneda |
  Imp. Neto Gravado | Imp. Neto No Gravado | Imp. Op. Exentas |
  IVA | Imp. Total
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path


@dataclass
class ComprobanteARCA:
    """Un comprobante parseado del archivo ARCA."""
    fecha: date
    tipo_comprobante: str
    punto_venta: int
    numero_desde: int
    numero_hasta: int
    cod_autorizacion: str | None
    tipo_doc_receptor: str | None
    nro_doc_receptor: str | None
    denominacion_receptor: str | None
    moneda: str
    tipo_cambio: Decimal
    importe_total: Decimal

    @property
    def es_nota_credito(self) -> bool:
        t = self.tipo_comprobante.upper()
        return "NOTA DE CREDITO" in t or "NOTA DE CRÉDITO" in t

    @property
    def importe_con_signo(self) -> Decimal:
        """Notas de crédito restan; facturas y ND suman."""
        return -self.importe_total if self.es_nota_credito else self.importe_total


def _normalizar_encabezado(col: str) -> str:
    """Normaliza un nombre de columna quitando acentos, puntos y espacios extra."""
    col = col.strip().lower()
    reemplazos = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "ñ": "n", ".": "", "nº": "n",
    }
    for old, new in reemplazos.items():
        col = col.replace(old, new)
    return col


# Mapeo de columna normalizada -> campo interno.
# Orden: patrones más específicos primero para evitar matches parciales.
_COLUMNAS_MAP = [
    ("punto de venta", "punto_venta"),
    ("numero desde", "numero_desde"),
    ("numero hasta", "numero_hasta"),
    ("cod autorizacion", "cod_autorizacion"),
    ("tipo doc receptor", "tipo_doc_receptor"),
    ("nro doc receptor", "nro_doc_receptor"),
    ("denominacion receptor", "denominacion_receptor"),
    ("tipo cambio", "tipo_cambio"),
    ("imp total", "importe_total"),
    ("moneda", "moneda"),
    ("fecha", "fecha"),
    ("tipo", "tipo_comprobante"),  # Al final: solo matchea si no matcheó antes
]


def _parsear_fecha(valor: str) -> date:
    valor = valor.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(valor, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Formato de fecha no reconocido: {valor}")


def _parsear_decimal(valor: str) -> Decimal:
    valor = valor.strip()
    if not valor or valor == "-":
        return Decimal("0.00")
    # ARCA puede usar coma como decimal
    valor = valor.replace(".", "").replace(",", ".")
    return Decimal(valor).quantize(Decimal("0.01"))


def _parsear_int(valor: str) -> int:
    valor = valor.strip()
    if not valor or valor == "-":
        return 0
    # Quitar separadores de miles
    valor = valor.replace(".", "").replace(",", "")
    return int(valor)


def _resolver_columnas(encabezados: list[str]) -> dict[int, str]:
    """Mapea índice de columna -> campo interno."""
    mapping = {}
    campos_usados = set()
    for i, col in enumerate(encabezados):
        normalizado = _normalizar_encabezado(col)
        for patron, campo in _COLUMNAS_MAP:
            if campo not in campos_usados and patron in normalizado:
                mapping[i] = campo
                campos_usados.add(campo)
                break
    return mapping


def _fila_a_comprobante(fila: list[str], col_map: dict[int, str]) -> ComprobanteARCA | None:
    """Convierte una fila a ComprobanteARCA. Retorna None si la fila está vacía."""
    datos: dict = {}
    for i, campo in col_map.items():
        if i >= len(fila):
            continue
        datos[campo] = fila[i].strip() if fila[i] else ""

    if not datos.get("fecha") or not datos.get("importe_total"):
        return None

    try:
        return ComprobanteARCA(
            fecha=_parsear_fecha(datos["fecha"]),
            tipo_comprobante=datos.get("tipo_comprobante", ""),
            punto_venta=_parsear_int(datos.get("punto_venta", "0")),
            numero_desde=_parsear_int(datos.get("numero_desde", "0")),
            numero_hasta=_parsear_int(datos.get("numero_hasta", "0")),
            cod_autorizacion=datos.get("cod_autorizacion") or None,
            tipo_doc_receptor=datos.get("tipo_doc_receptor") or None,
            nro_doc_receptor=datos.get("nro_doc_receptor") or None,
            denominacion_receptor=datos.get("denominacion_receptor") or None,
            moneda=datos.get("moneda", "PES"),
            tipo_cambio=_parsear_decimal(datos.get("tipo_cambio", "1")),
            importe_total=_parsear_decimal(datos["importe_total"]),
        )
    except (ValueError, KeyError):
        return None


def parsear_csv(contenido: bytes, encoding: str = "utf-8") -> list[ComprobanteARCA]:
    """Parsea un CSV de Mis Comprobantes ARCA (separado por ; o ,)."""
    texto = contenido.decode(encoding, errors="replace")
    # Detectar separador buscando en las primeras líneas (la primera puede ser un título)
    lineas = texto.split("\n")
    sep = ","
    for linea in lineas[:5]:
        if ";" in linea:
            sep = ";"
            break

    reader = csv.reader(io.StringIO(texto), delimiter=sep)
    filas = list(reader)

    if not filas:
        return []

    # Buscar fila de encabezado (puede haber una fila título antes)
    encabezado_idx = 0
    for i, fila in enumerate(filas[:5]):
        texto_fila = " ".join(fila).lower()
        if "fecha" in texto_fila and ("tipo" in texto_fila or "punto" in texto_fila):
            encabezado_idx = i
            break

    col_map = _resolver_columnas(filas[encabezado_idx])
    if not col_map:
        raise ValueError("No se encontraron columnas reconocibles en el archivo CSV.")

    resultado = []
    for fila in filas[encabezado_idx + 1:]:
        if not any(c.strip() for c in fila):
            continue
        comp = _fila_a_comprobante(fila, col_map)
        if comp:
            resultado.append(comp)

    return resultado


def parsear_excel(contenido: bytes) -> list[ComprobanteARCA]:
    """Parsea un archivo Excel (.xlsx) de Mis Comprobantes ARCA."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active

    filas = []
    for row in ws.iter_rows(values_only=True):
        filas.append([str(c) if c is not None else "" for c in row])
    wb.close()

    if not filas:
        return []

    # Buscar fila de encabezado
    encabezado_idx = 0
    for i, fila in enumerate(filas[:5]):
        texto_fila = " ".join(fila).lower()
        if "fecha" in texto_fila and ("tipo" in texto_fila or "punto" in texto_fila):
            encabezado_idx = i
            break

    col_map = _resolver_columnas(filas[encabezado_idx])
    if not col_map:
        raise ValueError("No se encontraron columnas reconocibles en el archivo Excel.")

    resultado = []
    for fila in filas[encabezado_idx + 1:]:
        if not any(c.strip() for c in fila):
            continue
        comp = _fila_a_comprobante(fila, col_map)
        if comp:
            resultado.append(comp)

    return resultado


def parsear_archivo(ruta: Path | str, contenido: bytes | None = None) -> list[ComprobanteARCA]:
    """Parsea un archivo ARCA (auto-detecta formato por extensión)."""
    ruta = Path(ruta)
    if contenido is None:
        contenido = ruta.read_bytes()

    ext = ruta.suffix.lower()
    if ext == ".csv":
        # Intentar UTF-8, si falla probar latin-1
        try:
            return parsear_csv(contenido, "utf-8")
        except UnicodeDecodeError:
            return parsear_csv(contenido, "latin-1")
    elif ext in (".xlsx", ".xls"):
        return parsear_excel(contenido)
    else:
        raise ValueError(f"Formato no soportado: {ext}. Use CSV o Excel (.xlsx).")
