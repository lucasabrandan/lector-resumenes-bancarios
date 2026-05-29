"""Servicio de gestión de percepciones/retenciones IIBB (SIRCREB)."""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal

from sqlalchemy import func, extract
from sqlalchemy.orm import Session

from app.db.models import PercepcionIIBBDB
from app.parsers.sircreb import PercepcionIIBB, JURISDICCIONES


def guardar_percepciones_iibb(
    percepciones: list[PercepcionIIBB],
    archivo_origen: str,
    db: Session,
) -> int:
    """Persiste percepciones/retenciones IIBB parseadas. Retorna cantidad."""
    for p in percepciones:
        db.add(PercepcionIIBBDB(
            jurisdiccion=p.jurisdiccion,
            jurisdiccion_nombre=p.jurisdiccion_nombre,
            cuit_agente=p.cuit_agente,
            fecha=p.fecha,
            tipo=p.tipo.value,
            monto_sujeto=float(p.monto_sujeto),
            alicuota=float(p.alicuota) if p.alicuota is not None else None,
            monto_retenido=float(p.monto_retenido),
            regimen=p.regimen,
            tipo_comprobante=p.tipo_comprobante,
            letra_comprobante=p.letra_comprobante,
            numero_comprobante=p.numero_comprobante,
            archivo_origen=archivo_origen,
        ))
    db.commit()
    return len(percepciones)


def archivo_sircreb_ya_cargado(nombre_archivo: str, db: Session) -> int:
    """Retorna cantidad de registros ya cargados de ese archivo."""
    return (
        db.query(PercepcionIIBBDB)
        .filter(PercepcionIIBBDB.archivo_origen == nombre_archivo)
        .count()
    )


def eliminar_sircreb_por_archivo(nombre_archivo: str, db: Session) -> int:
    """Elimina registros de un archivo. Retorna cantidad eliminada."""
    count = (
        db.query(PercepcionIIBBDB)
        .filter(PercepcionIIBBDB.archivo_origen == nombre_archivo)
        .delete()
    )
    db.commit()
    return count


def listar_archivos_sircreb(db: Session) -> list[dict]:
    """Lista archivos SIRCREB cargados con resumen."""
    query = (
        db.query(
            PercepcionIIBBDB.archivo_origen,
            func.count(PercepcionIIBBDB.id).label("cantidad"),
            func.min(PercepcionIIBBDB.fecha).label("fecha_desde"),
            func.max(PercepcionIIBBDB.fecha).label("fecha_hasta"),
            func.sum(PercepcionIIBBDB.monto_retenido).label("total"),
        )
        .group_by(PercepcionIIBBDB.archivo_origen)
    )

    return [
        {
            "archivo": r.archivo_origen,
            "cantidad": r.cantidad,
            "fecha_desde": r.fecha_desde,
            "fecha_hasta": r.fecha_hasta,
            "total": Decimal(str(r.total)).quantize(Decimal("0.01")),
        }
        for r in query.all()
    ]


MESES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


@dataclass
class LineaSircreb:
    """Línea del reporte: una jurisdicción en un mes."""
    anio: int
    mes: int
    jurisdiccion: int
    jurisdiccion_nombre: str
    tipo: str
    cantidad: int
    total: Decimal

    @property
    def es_retencion(self) -> bool:
        return self.tipo == "RETENCION"

    @property
    def periodo(self) -> str:
        return f"{MESES[self.mes]} {self.anio}"


@dataclass
class ResumenMensualSircreb:
    """Resumen de un mes con percepciones/retenciones por jurisdicción."""
    anio: int
    mes: int
    lineas: list[LineaSircreb]

    @property
    def total_percepciones(self) -> Decimal:
        return sum((l.total for l in self.lineas if not l.es_retencion), Decimal("0"))

    @property
    def total_retenciones(self) -> Decimal:
        return sum((l.total for l in self.lineas if l.es_retencion), Decimal("0"))

    @property
    def total_general(self) -> Decimal:
        return sum((l.total for l in self.lineas), Decimal("0"))

    @property
    def periodo(self) -> str:
        return f"{MESES[self.mes]} {self.anio}"


def generar_reporte_sircreb(db: Session) -> list[ResumenMensualSircreb]:
    """Genera reporte mensual de percepciones/retenciones IIBB por jurisdicción."""
    query = (
        db.query(
            extract("year", PercepcionIIBBDB.fecha).label("anio"),
            extract("month", PercepcionIIBBDB.fecha).label("mes"),
            PercepcionIIBBDB.jurisdiccion,
            PercepcionIIBBDB.jurisdiccion_nombre,
            PercepcionIIBBDB.tipo,
            func.count(PercepcionIIBBDB.id).label("cantidad"),
            func.coalesce(func.sum(PercepcionIIBBDB.monto_retenido), 0).label("total"),
        )
    )

    rows = (
        query
        .group_by("anio", "mes", PercepcionIIBBDB.jurisdiccion,
                  PercepcionIIBBDB.jurisdiccion_nombre, PercepcionIIBBDB.tipo)
        .order_by("anio", "mes", PercepcionIIBBDB.jurisdiccion_nombre)
        .all()
    )

    meses_dict: dict[tuple[int, int], list[LineaSircreb]] = {}
    for r in rows:
        clave = (int(r.anio), int(r.mes))
        linea = LineaSircreb(
            anio=int(r.anio),
            mes=int(r.mes),
            jurisdiccion=r.jurisdiccion,
            jurisdiccion_nombre=r.jurisdiccion_nombre,
            tipo=r.tipo,
            cantidad=r.cantidad,
            total=Decimal(str(r.total)).quantize(Decimal("0.01")),
        )
        meses_dict.setdefault(clave, []).append(linea)

    return [
        ResumenMensualSircreb(anio=anio, mes=mes, lineas=lineas)
        for (anio, mes), lineas in sorted(meses_dict.items())
    ]


def resumen_totales_sircreb(reporte: list[ResumenMensualSircreb]) -> dict:
    """Totales globales del reporte SIRCREB."""
    total_percepciones = sum((m.total_percepciones for m in reporte), Decimal("0"))
    total_retenciones = sum((m.total_retenciones for m in reporte), Decimal("0"))
    cantidad = sum(sum(l.cantidad for l in m.lineas) for m in reporte)
    return {
        "total_percepciones": total_percepciones,
        "total_retenciones": total_retenciones,
        "total_general": total_percepciones + total_retenciones,
        "cantidad": cantidad,
    }


def exportar_sircreb_xlsx(reporte: list[ResumenMensualSircreb]) -> "BytesIO":
    """Genera un archivo XLSX con el detalle de percepciones/retenciones IIBB."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "SIRCREB IIBB"

    money_format = '#,##0.00'
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    total_font = Font(bold=True, size=11)
    mes_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    mes_font = Font(bold=True, size=11)
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    ws.merge_cells("A1:F1")
    ws["A1"] = "Percepciones y Retenciones IIBB (SIRCREB)"
    ws["A1"].font = title_font

    ws.merge_cells("A2:F2")
    ws["A2"] = "Para carga en SIFERE Web y DDJJ de Ingresos Brutos"
    ws["A2"].font = Font(italic=True, color="666666")

    headers = ["Periodo", "Jurisdiccion", "Tipo", "Cantidad", "Importe"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_num = 5
    gran_total = Decimal("0")

    for mes in reporte:
        ws.cell(row=row_num, column=1, value=mes.periodo).font = mes_font
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).fill = mes_fill
        row_num += 1

        for linea in mes.lineas:
            ws.cell(row=row_num, column=1, value="")
            ws.cell(row=row_num, column=2, value=linea.jurisdiccion_nombre)
            tipo_label = "Retención" if linea.es_retencion else "Percepción"
            ws.cell(row=row_num, column=3, value=tipo_label)
            ws.cell(row=row_num, column=4, value=linea.cantidad)
            c = ws.cell(row=row_num, column=5, value=float(linea.total))
            c.number_format = money_format
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1

        ws.cell(row=row_num, column=2, value="Subtotal").font = Font(bold=True, size=10)
        ws.cell(row=row_num, column=4, value=sum(l.cantidad for l in mes.lineas)).font = Font(bold=True, size=10)
        c = ws.cell(row=row_num, column=5, value=float(mes.total_general))
        c.number_format = money_format
        c.font = Font(bold=True, size=10)
        row_num += 1
        gran_total += mes.total_general

    row_num += 1
    ws.cell(row=row_num, column=1, value="TOTAL GENERAL").font = total_font
    for col in range(1, 6):
        ws.cell(row=row_num, column=col).fill = total_fill
    c = ws.cell(row=row_num, column=5, value=float(gran_total))
    c.number_format = money_format
    c.font = total_font
    c.fill = total_fill

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
