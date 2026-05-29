"""
Servicio de movimientos: orquesta parsing, persistencia y consultas.

Es el "caso de uso" central de la app. El front y la API llaman acá,
nunca al parser o a la DB directamente.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

from sqlalchemy import func, extract
from sqlalchemy.orm import Session

from app.db.models import ComprobanteDB, MovimientoDB, PercepcionIIBBDB
from app.domain.models import MovimientoBancario, SignoMovimiento, TipoMovimiento, Moneda
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.parsers.supervielle_pdf import ParserSupervielle
from app.parsers.mercadopago_pdf import ParserMercadoPago, ResumenMP, es_pdf_mercadopago


def archivo_ya_cargado(nombre_archivo: str, db: Session) -> int:
    """Devuelve la cantidad de movimientos ya cargados de este archivo, o 0."""
    return (
        db.query(func.count(MovimientoDB.id))
        .filter(MovimientoDB.archivo_origen == nombre_archivo)
        .scalar() or 0
    )


def procesar_pdf(ruta_pdf: Path, db: Session) -> tuple[int, ResumenMP | None]:
    """Parsea un PDF bancario y guarda los movimientos en la DB.

    Auto-detecta si es MercadoPago o Supervielle.

    Returns:
        (cantidad_movimientos, resumen_mp_o_None)
    """
    resumen_mp = None
    if es_pdf_mercadopago(ruta_pdf):
        parser = ParserMercadoPago()
        movimientos, resumen_mp = parser.parsear_con_resumen(ruta_pdf)
    else:
        parser = ParserSupervielle()
        movimientos = parser.parsear(ruta_pdf)

    nombre_archivo = ruta_pdf.name

    registros = []
    for m in movimientos:
        registro = MovimientoDB(
            banco=m.banco,
            cuenta=m.cuenta,
            archivo_origen=nombre_archivo,
            pagina_origen=m.pagina_origen,
            fecha=m.fecha,
            concepto=m.concepto,
            detalle_adicional=m.detalle_adicional,
            numero_operacion=m.numero_operacion,
            importe=float(m.importe),
            signo=m.signo.value,
            saldo_posterior=float(m.saldo_posterior) if m.saldo_posterior is not None else None,
            moneda=m.moneda.value,
            tipo=m.tipo.value,
        )
        registros.append(registro)

    db.add_all(registros)
    db.commit()
    return len(registros), resumen_mp


def limpiar_todos_los_datos(db: Session) -> dict[str, int]:
    """Elimina todos los movimientos, comprobantes y percepciones IIBB."""
    resultado = {}
    resultado["movimientos"] = db.query(MovimientoDB).delete(synchronize_session=False)
    resultado["comprobantes"] = db.query(ComprobanteDB).delete(synchronize_session=False)
    resultado["percepciones_iibb"] = db.query(PercepcionIIBBDB).delete(synchronize_session=False)
    db.commit()
    return resultado


def eliminar_por_archivo(nombre_archivo: str, db: Session) -> int:
    """Elimina todos los movimientos de un archivo. Devuelve cantidad eliminada."""
    cantidad = (
        db.query(MovimientoDB)
        .filter(MovimientoDB.archivo_origen == nombre_archivo)
        .delete(synchronize_session=False)
    )
    db.commit()
    return cantidad


def listar_archivos_cargados(db: Session) -> list[dict]:
    """Lista archivos procesados con stats."""
    rows = (
        db.query(
            MovimientoDB.archivo_origen,
            func.count(MovimientoDB.id).label("cantidad"),
            func.min(MovimientoDB.fecha).label("fecha_min"),
            func.max(MovimientoDB.fecha).label("fecha_max"),
        )
        .group_by(MovimientoDB.archivo_origen)
        .order_by(MovimientoDB.archivo_origen)
        .all()
    )
    return [
        {
            "archivo": r.archivo_origen,
            "cantidad": r.cantidad,
            "fecha_min": r.fecha_min,
            "fecha_max": r.fecha_max,
        }
        for r in rows
    ]


def listar_movimientos(
    db: Session,
    cuenta: str | None = None,
    tipo: str | None = None,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    buscar: str | None = None,
    limite: int = 100,
    offset: int = 0,
) -> list[MovimientoDB]:
    """Consulta movimientos con filtros opcionales."""
    query = db.query(MovimientoDB)

    if cuenta:
        query = query.filter(MovimientoDB.cuenta == cuenta)
    if tipo:
        query = query.filter(MovimientoDB.tipo == tipo)
    if fecha_desde:
        query = query.filter(MovimientoDB.fecha >= fecha_desde)
    if fecha_hasta:
        query = query.filter(MovimientoDB.fecha <= fecha_hasta)
    if buscar:
        query = query.filter(MovimientoDB.concepto.ilike(f"%{buscar}%"))

    return query.order_by(MovimientoDB.fecha, MovimientoDB.id).offset(offset).limit(limite).all()


def contar_movimientos_filtrados(
    db: Session,
    tipo: str | None = None,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    buscar: str | None = None,
) -> int:
    """Total de movimientos con filtros aplicados."""
    query = db.query(func.count(MovimientoDB.id))
    if tipo:
        query = query.filter(MovimientoDB.tipo == tipo)
    if fecha_desde:
        query = query.filter(MovimientoDB.fecha >= fecha_desde)
    if fecha_hasta:
        query = query.filter(MovimientoDB.fecha <= fecha_hasta)
    if buscar:
        query = query.filter(MovimientoDB.concepto.ilike(f"%{buscar}%"))
    return query.scalar() or 0


def contar_movimientos(db: Session) -> int:
    """Total de movimientos en la DB."""
    return db.query(func.count(MovimientoDB.id)).scalar() or 0


def obtener_cuentas(db: Session) -> list[str]:
    """Lista de cuentas únicas en la DB."""
    rows = db.query(MovimientoDB.cuenta).distinct().all()
    return [r[0] for r in rows]


def obtener_rango_fechas(db: Session) -> tuple[date | None, date | None]:
    """Fecha mínima y máxima de los movimientos."""
    resultado = db.query(
        func.min(MovimientoDB.fecha),
        func.max(MovimientoDB.fecha),
    ).first()
    return (resultado[0], resultado[1]) if resultado else (None, None)


def resumen_mercadopago(db: Session) -> dict | None:
    """Genera resumen de movimientos MercadoPago.

    Returns None si no hay movimientos MP.
    """
    base = db.query(MovimientoDB).filter(MovimientoDB.banco == "MERCADOPAGO")

    total = base.count()
    if total == 0:
        return None

    totales = (
        db.query(
            MovimientoDB.signo,
            func.sum(MovimientoDB.importe).label("total"),
            func.count(MovimientoDB.id).label("cantidad"),
        )
        .filter(MovimientoDB.banco == "MERCADOPAGO")
        .group_by(MovimientoDB.signo)
        .all()
    )

    entradas = 0.0
    salidas = 0.0
    cant_entradas = 0
    cant_salidas = 0
    for r in totales:
        if r.signo == "CREDITO":
            entradas = float(r.total)
            cant_entradas = r.cantidad
        else:
            salidas = float(r.total)
            cant_salidas = r.cantidad

    fechas = (
        db.query(
            func.min(MovimientoDB.fecha).label("desde"),
            func.max(MovimientoDB.fecha).label("hasta"),
        )
        .filter(MovimientoDB.banco == "MERCADOPAGO")
        .first()
    )

    por_tipo = (
        db.query(
            MovimientoDB.tipo,
            MovimientoDB.signo,
            func.sum(MovimientoDB.importe).label("total"),
            func.count(MovimientoDB.id).label("cantidad"),
        )
        .filter(MovimientoDB.banco == "MERCADOPAGO")
        .group_by(MovimientoDB.tipo, MovimientoDB.signo)
        .order_by(func.sum(MovimientoDB.importe).desc())
        .all()
    )

    mensual = (
        db.query(
            extract("year", MovimientoDB.fecha).label("anio"),
            extract("month", MovimientoDB.fecha).label("mes"),
            MovimientoDB.signo,
            func.sum(MovimientoDB.importe).label("total"),
        )
        .filter(MovimientoDB.banco == "MERCADOPAGO")
        .group_by("anio", "mes", MovimientoDB.signo)
        .order_by("anio", "mes")
        .all()
    )

    meses_data = {}
    for r in mensual:
        clave = f"{int(r.anio)}-{int(r.mes):02d}"
        if clave not in meses_data:
            meses_data[clave] = {"entradas": 0.0, "salidas": 0.0}
        if r.signo == "CREDITO":
            meses_data[clave]["entradas"] = float(r.total)
        else:
            meses_data[clave]["salidas"] = float(r.total)

    return {
        "total": total,
        "entradas": entradas,
        "salidas": salidas,
        "cant_entradas": cant_entradas,
        "cant_salidas": cant_salidas,
        "neto": entradas - salidas,
        "fecha_desde": fechas.desde if fechas else None,
        "fecha_hasta": fechas.hasta if fechas else None,
        "por_tipo": [
            {"tipo": r.tipo, "signo": r.signo, "total": float(r.total), "cantidad": r.cantidad}
            for r in por_tipo
        ],
        "meses": meses_data,
    }


def exportar_movimientos_xlsx(movimientos: list[MovimientoDB]) -> BytesIO:
    """Genera un XLSX con el listado de movimientos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    money_format = '#,##0.00'
    thin_border = Border(
        bottom=Side(style="thin", color="E5E7EB"),
    )

    headers = ["Fecha", "Banco", "Cuenta", "Concepto", "Detalle", "Tipo", "Signo", "Importe", "Saldo", "ID Operación", "Archivo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, m in enumerate(movimientos, 2):
        ws.cell(row=i, column=1, value=m.fecha.isoformat() if m.fecha else "")
        ws.cell(row=i, column=2, value=m.banco)
        ws.cell(row=i, column=3, value=m.cuenta)
        ws.cell(row=i, column=4, value=m.concepto)
        ws.cell(row=i, column=5, value=m.detalle_adicional or "")
        ws.cell(row=i, column=6, value=m.tipo)
        ws.cell(row=i, column=7, value=m.signo)
        ws.cell(row=i, column=8, value=float(m.importe)).number_format = money_format
        ws.cell(row=i, column=9, value=float(m.saldo_posterior) if m.saldo_posterior is not None else "").number_format = money_format
        ws.cell(row=i, column=10, value=m.numero_operacion or "")
        ws.cell(row=i, column=11, value=m.archivo_origen or "")
        for col in range(1, 12):
            ws.cell(row=i, column=col).border = thin_border

    if movimientos:
        total_row = len(movimientos) + 2
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        total_debitos = sum(float(m.importe) for m in movimientos if m.signo == "DEBITO")
        total_creditos = sum(float(m.importe) for m in movimientos if m.signo == "CREDITO")

        ws.cell(row=total_row, column=6, value="DÉBITOS:").font = total_font
        ws.cell(row=total_row, column=7, value="DEBITO").font = total_font
        ws.cell(row=total_row, column=8, value=total_debitos).number_format = money_format
        ws.cell(row=total_row, column=8).font = total_font

        ws.cell(row=total_row + 1, column=6, value="CRÉDITOS:").font = total_font
        ws.cell(row=total_row + 1, column=7, value="CREDITO").font = total_font
        ws.cell(row=total_row + 1, column=8, value=total_creditos).number_format = money_format
        ws.cell(row=total_row + 1, column=8).font = total_font

        ws.cell(row=total_row + 2, column=6, value="NETO:").font = total_font
        ws.cell(row=total_row + 2, column=8, value=total_creditos - total_debitos).number_format = money_format
        ws.cell(row=total_row + 2, column=8).font = total_font

        for r in range(total_row, total_row + 3):
            for col in range(1, 12):
                ws.cell(row=r, column=col).fill = total_fill

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
