"""
Reporte de percepciones y retenciones sufridas.

Agrupa por mes y tipo (Percepcion IVA, Retencion IIBB, etc.) los montos
que el banco le aplico al cliente. La contadora usa este reporte para:
  - Cargar las percepciones/retenciones en el SIRE.
  - Tomarse el credito fiscal correspondiente en la DDJJ de IVA.
  - Presentar ante ARCA.
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO

from sqlalchemy import func, extract, case
from sqlalchemy.orm import Session

from app.db.models import MovimientoDB


# Tipos de movimiento que son percepciones o retenciones
TIPOS_PERCEPCIONES_RETENCIONES = [
    "PERCEPCION_IVA",
    "PERCEPCION_IIBB",
    "PERCEPCION_GANANCIAS",
    "RETENCION_IVA",
    "RETENCION_IIBB",
    "RETENCION_GANANCIAS",
    "RETENCION_SUSS",
]

# Etiquetas amigables para cada tipo
ETIQUETAS_TIPO = {
    "PERCEPCION_IVA": "Percepcion IVA",
    "PERCEPCION_IIBB": "Percepcion IIBB",
    "PERCEPCION_GANANCIAS": "Percepcion Ganancias",
    "RETENCION_IVA": "Retencion IVA",
    "RETENCION_IIBB": "Retencion IIBB",
    "RETENCION_GANANCIAS": "Retencion Ganancias",
    "RETENCION_SUSS": "Retencion SUSS",
}

MESES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


@dataclass
class LineaPercepcion:
    """Una linea del reporte: un tipo de percepcion/retencion en un mes."""
    anio: int
    mes: int
    tipo: str
    cantidad: int
    total: Decimal

    @property
    def etiqueta(self) -> str:
        return ETIQUETAS_TIPO.get(self.tipo, self.tipo)

    @property
    def es_retencion(self) -> bool:
        return self.tipo.startswith("RETENCION")

    @property
    def mes_nombre(self) -> str:
        return MESES[self.mes]

    @property
    def periodo(self) -> str:
        return f"{self.mes_nombre} {self.anio}"


@dataclass
class ResumenMensual:
    """Resumen de un mes con todas las percepciones/retenciones."""
    anio: int
    mes: int
    lineas: list[LineaPercepcion]

    @property
    def total_percepciones(self) -> Decimal:
        return sum((l.total for l in self.lineas if not l.es_retencion), Decimal("0"))

    @property
    def total_retenciones(self) -> Decimal:
        return sum((l.total for l in self.lineas if l.es_retencion), Decimal("0"))

    @property
    def total_general(self) -> Decimal:
        return sum((l.total for l in self.lineas), Decimal("0"))

    @property
    def mes_nombre(self) -> str:
        return MESES[self.mes]

    @property
    def periodo(self) -> str:
        return f"{self.mes_nombre} {self.anio}"


def generar_reporte_percepciones(
    db: Session, cliente_ids: list[int] | None = None
) -> list[ResumenMensual]:
    """Genera el reporte mensual de percepciones y retenciones.

    Returns:
        Lista ordenada cronologicamente, cada elemento con sus lineas por tipo.
    """
    from app.services.movimientos import _aplicar_filtro_clientes

    query = (
        db.query(
            extract("year", MovimientoDB.fecha).label("anio"),
            extract("month", MovimientoDB.fecha).label("mes"),
            MovimientoDB.tipo,
            func.count(MovimientoDB.id).label("cantidad"),
            func.coalesce(func.sum(MovimientoDB.importe), 0).label("total"),
        )
        .filter(MovimientoDB.tipo.in_(TIPOS_PERCEPCIONES_RETENCIONES))
    )
    query = _aplicar_filtro_clientes(query, cliente_ids)
    rows = (
        query
        .group_by("anio", "mes", MovimientoDB.tipo)
        .order_by("anio", "mes", MovimientoDB.tipo)
        .all()
    )

    # Agrupar por mes
    meses_dict: dict[tuple[int, int], list[LineaPercepcion]] = {}
    for r in rows:
        clave = (int(r.anio), int(r.mes))
        linea = LineaPercepcion(
            anio=int(r.anio),
            mes=int(r.mes),
            tipo=r.tipo,
            cantidad=r.cantidad,
            total=Decimal(str(r.total)).quantize(Decimal("0.01")),
        )
        meses_dict.setdefault(clave, []).append(linea)

    return [
        ResumenMensual(anio=anio, mes=mes, lineas=lineas)
        for (anio, mes), lineas in sorted(meses_dict.items())
    ]


def resumen_totales(reporte: list[ResumenMensual]) -> dict:
    """Totales globales del reporte para mostrar en el dashboard."""
    total_percepciones = sum((m.total_percepciones for m in reporte), Decimal("0"))
    total_retenciones = sum((m.total_retenciones for m in reporte), Decimal("0"))
    cantidad = sum(sum(l.cantidad for l in m.lineas) for m in reporte)
    return {
        "total_percepciones": total_percepciones,
        "total_retenciones": total_retenciones,
        "total_general": total_percepciones + total_retenciones,
        "cantidad": cantidad,
    }


def exportar_percepciones_xlsx(reporte: list[ResumenMensual]) -> BytesIO:
    """Genera un archivo XLSX con el detalle de percepciones y retenciones."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Percepciones y Retenciones"

    money_format = '#,##0.00'
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    total_font = Font(bold=True, size=11)
    mes_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    mes_font = Font(bold=True, size=11)
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    # Titulo
    ws.merge_cells("A1:E1")
    ws["A1"] = "Reporte de Percepciones y Retenciones Sufridas"
    ws["A1"].font = title_font

    ws.merge_cells("A2:E2")
    ws["A2"] = "Para carga en SIRE y DDJJ de IVA"
    ws["A2"].font = Font(italic=True, color="666666")

    # Headers
    headers = ["Periodo", "Tipo", "Cantidad", "Importe"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_num = 5
    gran_total = Decimal("0")

    for mes in reporte:
        # Fila de mes
        ws.cell(row=row_num, column=1, value=mes.periodo).font = mes_font
        for col in range(1, 5):
            ws.cell(row=row_num, column=col).fill = mes_fill
        row_num += 1

        for linea in mes.lineas:
            ws.cell(row=row_num, column=1, value="")
            ws.cell(row=row_num, column=2, value=linea.etiqueta)
            ws.cell(row=row_num, column=3, value=linea.cantidad)
            c = ws.cell(row=row_num, column=4, value=float(linea.total))
            c.number_format = money_format
            for col in range(1, 5):
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1

        # Subtotal del mes
        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value="Subtotal").font = Font(bold=True, size=10)
        ws.cell(row=row_num, column=3, value=sum(l.cantidad for l in mes.lineas)).font = Font(bold=True, size=10)
        c = ws.cell(row=row_num, column=4, value=float(mes.total_general))
        c.number_format = money_format
        c.font = Font(bold=True, size=10)
        row_num += 1

        gran_total += mes.total_general

    # Fila de total general
    row_num += 1
    ws.cell(row=row_num, column=1, value="TOTAL GENERAL").font = total_font
    ws.cell(row=row_num, column=1).fill = total_fill
    for col in range(2, 5):
        ws.cell(row=row_num, column=col).fill = total_fill
    c = ws.cell(row=row_num, column=4, value=float(gran_total))
    c.number_format = money_format
    c.font = total_font
    c.fill = total_fill

    # Anchos
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
