"""
Reporte de Impuesto Ley 25.413 (débitos y créditos bancarios).

Este es EL reporte que la contadora necesita para presentar a ARCA.
Totaliza por mes cuánto cobró el banco en concepto de:
  - Impuesto sobre débitos (IMPUESTO_LEY_25413_SOBRE_DEBITOS)
  - Impuesto sobre créditos (IMPUESTO_LEY_25413_SOBRE_CREDITOS)

Nota sobre devoluciones (descubierto en iter 3.3):
    El banco a veces devuelve parte del impuesto por reclamos. Esas
    devoluciones se clasifican como DEVOLUCION y se restan del total
    del mes para obtener el impuesto neto.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from io import BytesIO

from sqlalchemy import func, extract, case
from sqlalchemy.orm import Session

from app.db.models import MovimientoDB


@dataclass
class TotalMensualLey25413:
    """Totales del impuesto para un mes específico."""
    anio: int
    mes: int
    impuesto_sobre_debitos: Decimal
    impuesto_sobre_creditos: Decimal
    devoluciones_debitos: Decimal = field(default_factory=lambda: Decimal("0.00"))

    @property
    def neto_sobre_debitos(self) -> Decimal:
        """Impuesto neto = bruto - devoluciones imputables."""
        return self.impuesto_sobre_debitos - self.devoluciones_debitos

    @property
    def total_bruto(self) -> Decimal:
        return self.impuesto_sobre_debitos + self.impuesto_sobre_creditos

    @property
    def total_neto(self) -> Decimal:
        return self.neto_sobre_debitos + self.impuesto_sobre_creditos

    @property
    def mes_nombre(self) -> str:
        meses = [
            "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
        ]
        return meses[self.mes]

    @property
    def periodo(self) -> str:
        return f"{self.mes_nombre} {self.anio}"


def generar_reporte_ley_25413(db: Session) -> list[TotalMensualLey25413]:
    """Genera el reporte mensual de Ley 25.413 con devoluciones.

    Returns:
        Lista ordenada cronológicamente con los totales por mes.
    """
    rows = (
        db.query(
            extract("year", MovimientoDB.fecha).label("anio"),
            extract("month", MovimientoDB.fecha).label("mes"),
            func.coalesce(
                func.sum(
                    case(
                        (MovimientoDB.tipo == "IMPUESTO_LEY_25413_SOBRE_DEBITOS", MovimientoDB.importe),
                        else_=0,
                    )
                ),
                0,
            ).label("imp_debitos"),
            func.coalesce(
                func.sum(
                    case(
                        (MovimientoDB.tipo == "IMPUESTO_LEY_25413_SOBRE_CREDITOS", MovimientoDB.importe),
                        else_=0,
                    )
                ),
                0,
            ).label("imp_creditos"),
            func.coalesce(
                func.sum(
                    case(
                        (
                            (MovimientoDB.tipo == "DEVOLUCION")
                            & (MovimientoDB.concepto.ilike("%imp%débit%")),
                            MovimientoDB.importe,
                        ),
                        else_=0,
                    )
                ),
                0,
            ).label("devoluciones_db"),
        )
        .filter(
            MovimientoDB.tipo.in_([
                "IMPUESTO_LEY_25413_SOBRE_DEBITOS",
                "IMPUESTO_LEY_25413_SOBRE_CREDITOS",
                "DEVOLUCION",
            ])
        )
        .group_by("anio", "mes")
        .order_by("anio", "mes")
        .all()
    )

    return [
        TotalMensualLey25413(
            anio=int(r.anio),
            mes=int(r.mes),
            impuesto_sobre_debitos=Decimal(str(r.imp_debitos)).quantize(Decimal("0.01")),
            impuesto_sobre_creditos=Decimal(str(r.imp_creditos)).quantize(Decimal("0.01")),
            devoluciones_debitos=Decimal(str(r.devoluciones_db)).quantize(Decimal("0.01")),
        )
        for r in rows
    ]


def resumen_general(db: Session) -> dict:
    """Estadísticas generales para el dashboard."""
    from app.services.movimientos import contar_movimientos, obtener_rango_fechas

    total_movs = contar_movimientos(db)
    fecha_min, fecha_max = obtener_rango_fechas(db)

    totales_signo = (
        db.query(
            MovimientoDB.signo,
            func.sum(MovimientoDB.importe).label("total"),
            func.count(MovimientoDB.id).label("cantidad"),
        )
        .group_by(MovimientoDB.signo)
        .all()
    )

    debitos = next((r for r in totales_signo if r.signo == "DEBITO"), None)
    creditos = next((r for r in totales_signo if r.signo == "CREDITO"), None)

    return {
        "total_movimientos": total_movs,
        "fecha_desde": fecha_min,
        "fecha_hasta": fecha_max,
        "total_debitos": Decimal(str(debitos.total)).quantize(Decimal("0.01")) if debitos else Decimal("0"),
        "cant_debitos": debitos.cantidad if debitos else 0,
        "total_creditos": Decimal(str(creditos.total)).quantize(Decimal("0.01")) if creditos else Decimal("0"),
        "cant_creditos": creditos.cantidad if creditos else 0,
    }


def exportar_reporte_xlsx(reporte: list[TotalMensualLey25413]) -> BytesIO:
    """Genera un archivo XLSX con el reporte Ley 25.413.

    Returns:
        BytesIO con el contenido del archivo listo para descargar.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, numbers, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Ley 25413"

    # Estilos
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    money_format = '#,##0.00'
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    total_font = Font(bold=True, size=11)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Título
    ws.merge_cells("A1:F1")
    ws["A1"] = "Reporte Impuesto Ley 25.413 — Débitos y Créditos Bancarios"
    ws["A1"].font = title_font

    ws.merge_cells("A2:F2")
    ws["A2"] = "Generado para presentación ante ARCA"
    ws["A2"].font = Font(italic=True, color="666666")

    # Headers (fila 4)
    headers = [
        "Período",
        "Imp. s/Débitos (Bruto)",
        "Devoluciones Imp. Débitos",
        "Imp. s/Débitos (Neto)",
        "Imp. s/Créditos",
        "Total Neto del Mes",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Datos
    total_db_bruto = Decimal("0")
    total_devoluciones = Decimal("0")
    total_db_neto = Decimal("0")
    total_cr = Decimal("0")
    total_neto = Decimal("0")

    for i, r in enumerate(reporte):
        row = 5 + i
        ws.cell(row=row, column=1, value=r.periodo)
        ws.cell(row=row, column=2, value=float(r.impuesto_sobre_debitos)).number_format = money_format
        ws.cell(row=row, column=3, value=float(r.devoluciones_debitos)).number_format = money_format
        ws.cell(row=row, column=4, value=float(r.neto_sobre_debitos)).number_format = money_format
        ws.cell(row=row, column=5, value=float(r.impuesto_sobre_creditos)).number_format = money_format
        ws.cell(row=row, column=6, value=float(r.total_neto)).number_format = money_format

        # Borde sutil entre filas
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border

        total_db_bruto += r.impuesto_sobre_debitos
        total_devoluciones += r.devoluciones_debitos
        total_db_neto += r.neto_sobre_debitos
        total_cr += r.impuesto_sobre_creditos
        total_neto += r.total_neto

    # Fila de totales
    total_row = 5 + len(reporte)
    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=2, value=float(total_db_bruto)).number_format = money_format
    ws.cell(row=total_row, column=3, value=float(total_devoluciones)).number_format = money_format
    ws.cell(row=total_row, column=4, value=float(total_db_neto)).number_format = money_format
    ws.cell(row=total_row, column=5, value=float(total_cr)).number_format = money_format
    ws.cell(row=total_row, column=6, value=float(total_neto)).number_format = money_format

    for col in range(1, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = total_fill
        cell.font = total_font

    # Anchos de columna
    ws.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
