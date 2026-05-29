"""
Reporte de conceptos recurrentes: agrupa movimientos por concepto y mes.

Permite a la contadora ver qué conceptos se repiten mes a mes,
cuántas veces aparecen y el detalle de importes.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import func, extract, case
from sqlalchemy.orm import Session

from app.db.models import MovimientoDB


@dataclass
class MesDetalle:
    """Datos de un concepto en un mes específico."""
    cantidad: int = 0
    total_debito: Decimal = Decimal("0")
    total_credito: Decimal = Decimal("0")

    @property
    def total_neto(self) -> Decimal:
        return self.total_credito - self.total_debito


@dataclass
class ConceptoRecurrente:
    """Un concepto con su desglose mensual."""
    concepto: str
    tipo: str
    meses: dict[str, MesDetalle] = field(default_factory=dict)
    total_cantidad: int = 0
    total_debito: Decimal = Decimal("0")
    total_credito: Decimal = Decimal("0")

    @property
    def total_neto(self) -> Decimal:
        return self.total_credito - self.total_debito

    @property
    def cantidad_meses(self) -> int:
        return len(self.meses)


@dataclass
class ReporteConceptos:
    """Resultado completo del reporte."""
    conceptos: list[ConceptoRecurrente]
    periodos: list[str]  # ["2024-01", "2024-02", ...]
    total_cantidad: int = 0
    total_debito: Decimal = Decimal("0")
    total_credito: Decimal = Decimal("0")

    @property
    def total_neto(self) -> Decimal:
        return self.total_credito - self.total_debito


def generar_reporte_conceptos(
    db: Session,
    tipo: str | None = None,
    signo: str | None = None,
    buscar: str | None = None,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
) -> ReporteConceptos:
    """Genera el reporte de conceptos recurrentes agrupados por mes."""

    query = db.query(
        MovimientoDB.concepto,
        MovimientoDB.tipo,
        extract("year", MovimientoDB.fecha).label("anio"),
        extract("month", MovimientoDB.fecha).label("mes"),
        func.count(MovimientoDB.id).label("cantidad"),
        func.sum(
            case(
                (MovimientoDB.signo == "DEBITO", MovimientoDB.importe),
                else_=0,
            )
        ).label("total_debito"),
        func.sum(
            case(
                (MovimientoDB.signo == "CREDITO", MovimientoDB.importe),
                else_=0,
            )
        ).label("total_credito"),
    )

    if tipo:
        query = query.filter(MovimientoDB.tipo == tipo)
    if signo:
        query = query.filter(MovimientoDB.signo == signo)
    if buscar:
        query = query.filter(MovimientoDB.concepto.ilike(f"%{buscar}%"))
    if fecha_desde:
        query = query.filter(MovimientoDB.fecha >= fecha_desde)
    if fecha_hasta:
        query = query.filter(MovimientoDB.fecha <= fecha_hasta)

    rows = (
        query
        .group_by(MovimientoDB.concepto, MovimientoDB.tipo, "anio", "mes")
        .order_by(MovimientoDB.concepto, "anio", "mes")
        .all()
    )

    # Construir estructura
    conceptos_map: dict[str, ConceptoRecurrente] = {}
    periodos_set: set[str] = set()

    for r in rows:
        periodo = f"{int(r.anio)}-{int(r.mes):02d}"
        periodos_set.add(periodo)

        key = r.concepto
        if key not in conceptos_map:
            conceptos_map[key] = ConceptoRecurrente(
                concepto=r.concepto,
                tipo=r.tipo,
            )

        cr = conceptos_map[key]
        detalle = MesDetalle(
            cantidad=r.cantidad,
            total_debito=Decimal(str(r.total_debito)),
            total_credito=Decimal(str(r.total_credito)),
        )
        cr.meses[periodo] = detalle
        cr.total_cantidad += r.cantidad
        cr.total_debito += detalle.total_debito
        cr.total_credito += detalle.total_credito

    periodos = sorted(periodos_set)
    conceptos = sorted(conceptos_map.values(), key=lambda c: c.total_cantidad, reverse=True)

    total_cantidad = sum(c.total_cantidad for c in conceptos)
    total_debito = sum(c.total_debito for c in conceptos)
    total_credito = sum(c.total_credito for c in conceptos)

    return ReporteConceptos(
        conceptos=conceptos,
        periodos=periodos,
        total_cantidad=total_cantidad,
        total_debito=total_debito,
        total_credito=total_credito,
    )


MESES_NOMBRE = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
}


def periodo_label(periodo: str) -> str:
    """Convierte '2024-01' en 'Ene 2024'."""
    anio, mes = periodo.split("-")
    return f"{MESES_NOMBRE[int(mes)]} {anio}"


def exportar_conceptos_xlsx(reporte: ReporteConceptos) -> BytesIO:
    """Exporta el reporte a XLSX con subtotales y totales."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Conceptos Recurrentes"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    subtotal_font = Font(bold=True, size=10)
    subtotal_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    money_fmt = '#,##0.00'
    thin_border = Border(bottom=Side(style="thin", color="E5E7EB"))

    # Headers
    headers = ["Concepto", "Tipo", "Meses"]
    for p in reporte.periodos:
        headers.append(f"{periodo_label(p)} (Cant)")
        headers.append(f"{periodo_label(p)} (Debito)")
        headers.append(f"{periodo_label(p)} (Credito)")
    headers += ["Total Cant", "Total Debito", "Total Credito", "Neto"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for cr in reporte.conceptos:
        ws.cell(row=row, column=1, value=cr.concepto)
        ws.cell(row=row, column=2, value=cr.tipo.replace("_", " "))
        ws.cell(row=row, column=3, value=cr.cantidad_meses)

        col = 4
        for p in reporte.periodos:
            det = cr.meses.get(p)
            if det:
                ws.cell(row=row, column=col, value=det.cantidad)
                ws.cell(row=row, column=col + 1, value=float(det.total_debito)).number_format = money_fmt
                ws.cell(row=row, column=col + 2, value=float(det.total_credito)).number_format = money_fmt
            col += 3

        ws.cell(row=row, column=col, value=cr.total_cantidad)
        ws.cell(row=row, column=col + 1, value=float(cr.total_debito)).number_format = money_fmt
        ws.cell(row=row, column=col + 2, value=float(cr.total_credito)).number_format = money_fmt
        ws.cell(row=row, column=col + 3, value=float(cr.total_neto)).number_format = money_fmt

        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = thin_border

        row += 1

    # Fila de totales
    if reporte.conceptos:
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).font = total_font
            ws.cell(row=row, column=c).fill = total_fill

        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=3, value=len(reporte.periodos))

        col = 4
        for p in reporte.periodos:
            cant = sum(cr.meses[p].cantidad for cr in reporte.conceptos if p in cr.meses)
            deb = sum(cr.meses[p].total_debito for cr in reporte.conceptos if p in cr.meses)
            cre = sum(cr.meses[p].total_credito for cr in reporte.conceptos if p in cr.meses)
            ws.cell(row=row, column=col, value=cant)
            ws.cell(row=row, column=col + 1, value=float(deb)).number_format = money_fmt
            ws.cell(row=row, column=col + 2, value=float(cre)).number_format = money_fmt
            col += 3

        ws.cell(row=row, column=col, value=reporte.total_cantidad)
        ws.cell(row=row, column=col + 1, value=float(reporte.total_debito)).number_format = money_fmt
        ws.cell(row=row, column=col + 2, value=float(reporte.total_credito)).number_format = money_fmt
        ws.cell(row=row, column=col + 3, value=float(reporte.total_neto)).number_format = money_fmt

    # Column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 8

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
